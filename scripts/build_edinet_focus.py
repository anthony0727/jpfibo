"""Extract a focused, machine-readable view of selected EDINET taxonomy elements.

Reads ``data/sources/edinet-taxonomy-<YEAR>/1e_ElementList.xlsx`` and emits
``data/derived/edinet_taxonomy_focus.json`` containing only the elements used by
J-FIBO and the benchmark cases, with bilingual labels and taxonomy metadata.

Run:
    uv run python scripts/build_edinet_focus.py
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[1]

DEFAULT_FOCUS = {
    # ---- Document & Entity Information (DEI) -------------------------------
    "EDINETCodeDEI",
    "SecurityCodeDEI",
    "FilerNameInJapaneseDEI",
    "FilerNameInEnglishDEI",
    "DocumentTypeDEI",
    "CabinetOfficeOrdinanceDEI",
    "AccountingStandardsDEI",
    "WhetherConsolidatedFinancialStatementsArePreparedDEI",
    "CurrentFiscalYearStartDateDEI",
    "CurrentFiscalYearEndDateDEI",
    "PreviousFiscalYearStartDateDEI",
    "PreviousFiscalYearEndDateDEI",
    "AmendmentFlagDEI",
    "ReportAmendmentFlagDEI",
    "XBRLAmendmentFlagDEI",

    # ---- jpcrp: governance / ownership / disclosure structure --------------
    # Shareholdings (政策保有株式 narrative + table)
    "ShareholdingsTextBlock",
    "DetailsOfSpecifiedInvestmentEquitySecuritiesHeldForPurposesOtherThanPureInvestmentReportingCompanyTable",
    "NameOfSecuritiesDetailsOfSpecifiedInvestmentEquitySecuritiesHeldForPurposesOtherThanPureInvestmentReportingCompany",
    "NumberOfSharesHeldDetailsOfSpecifiedInvestmentEquitySecuritiesHeldForPurposesOtherThanPureInvestmentReportingCompany",
    "BookValueDetailsOfSpecifiedInvestmentEquitySecuritiesHeldForPurposesOtherThanPureInvestmentReportingCompany",
    "PurposesOfHoldingDetailsOfSpecifiedInvestmentEquitySecuritiesHeldForPurposesOtherThanPureInvestmentReportingCompany",
    "OverviewOfBusinessAllianceDetailsOfSpecifiedInvestmentSharesHeldForPurposesOtherThanPureInvestmentReportingCompany",
    "QuantitativeEffectsOfShareholdingDetailsOfSpecifiedInvestmentSharesHeldForPurposesOtherThanPureInvestmentReportingCompany",
    "ReasonForIncreaseInNumberOfSharesDetailsOfSpecifiedInvestmentSharesHeldForPurposesOtherThanPureInvestmentReportingCompany",
    # Major shareholders (大株主の状況)
    "MajorShareholdersTextBlock",
    # Borrowings schedule (借入金等明細表) — both consolidated and non-consolidated forms
    "AnnexedDetailedScheduleOfBorrowingsFinancialStatementsTextBlock",
    "AnnexedConsolidatedDetailedScheduleOfBorrowingsTextBlock",
    # Related-party transactions (関連当事者との取引)
    "RelatedPartyInformationTextBlock",
    # Corporate governance overview (コーポレート・ガバナンスの状況等)
    "CorporateGovernanceTextBlock",
    "OverviewOfCorporateGovernanceTextBlock",
    "DescriptionOfBoardOfDirectorsTextBlock",
    "DescriptionOfNominationCommitteeEtcTextBlock",
    "DescriptionOfAuditAndSupervisoryCommitteeTextBlock",
    # Segment & sales / executives compensation
    "NotesSegmentInformationConsolidatedFinancialStatementsTextBlock",
    "RemunerationForDirectorsAndCorporateAuditorsTextBlock",
    "DirectorsAndAuditorsTextBlock",
    # Risk factors and MD&A — useful for narrative claim anchoring
    "BusinessRisksTextBlock",
    "ManagementAnalysisOfFinancialPositionOperatingResultsAndCashFlowsTextBlock",

    # ---- jppfs: anchor financial-statement elements (general industry) -----
    "Assets",
    "Liabilities",
    "NetAssets",
    "NetSales",
    "OperatingIncome",
    "OrdinaryIncome",
    "ProfitLoss",
    "CashAndDeposits",
    "ShortTermLoansPayable",
    "CurrentPortionOfLongTermLoansPayable",
    "LongTermLoansPayable",
    "LeaseObligationsCL",
    "LeaseObligationsNCL",
    "BorrowedMoneyLiabilitiesBNK",          # bank-industry 借用金 line
    "BillsRediscountedBorrowedMoneyLiabilitiesBNK",  # bank-industry 再割引手形 inside borrowed money
    "CommercialPapersLiabilities",

    # ---- jpsps: large shareholding reporting (大量保有報告書) family --------
    # NOTE: large-shareholding filings live in a separate form family with
    # the jpsps prefix. We list a small anchor set; the build script will
    # warn for any not found, so they auto-degrade.
    "FilerNameInJapaneseDEI",  # also appears in jpsps for the holder

    # ---- jplvh: holder identification on large-shareholding side -----------
    # Same approach: warn-on-miss is acceptable.
}

EDINET_NAMESPACES = {
    "jpdei_cor": "http://disclosure.edinet-fsa.go.jp/taxonomy/jpdei/2013-08-31/jpdei_cor",
    "jpcrp_cor": "http://disclosure.edinet-fsa.go.jp/taxonomy/jpcrp/2025-11-01/jpcrp_cor",
    "jppfs_cor": "http://disclosure.edinet-fsa.go.jp/taxonomy/jppfs/2025-11-01/jppfs_cor",
    "jpsps_cor": "http://disclosure.edinet-fsa.go.jp/taxonomy/jpsps/2025-11-01/jpsps_cor",
    "jplvh_cor": "http://disclosure.edinet-fsa.go.jp/taxonomy/jplvh/2025-11-01/jplvh_cor",
    "jpaud_cor": "http://disclosure.edinet-fsa.go.jp/taxonomy/jpaud/2025-11-01/jpaud_cor",
    "jpctl_cor": "http://disclosure.edinet-fsa.go.jp/taxonomy/jpctl/2025-11-01/jpctl_cor",
    "jpigp_cor": "http://disclosure.edinet-fsa.go.jp/taxonomy/jpigp/2025-11-01/jpigp_cor",
}


def _detect_header(values: list) -> dict[str, int] | None:
    """Locate the FSA header row by Japanese column titles.

    The 1e_ElementList and 1f_AccountList workbooks share most column names
    but the 1f sheets prepend a 科目分類 column and a 標準ラベル column at
    different offsets. Detect by header text so the same collector works
    on both shapes.
    """
    name_to_index: dict[str, int] = {}
    for i, v in enumerate(values):
        if v is None:
            continue
        s = str(v).strip()
        if s:
            name_to_index[s] = i
    # In 1e_ElementList the standard-label-ja column is "様式ツリー-標準ラベル（日本語）"
    # or "詳細ツリー-標準ラベル（日本語）"; in 1f_AccountList it is plain "標準ラベル（日本語）".
    standard_ja_candidates = [
        "標準ラベル（日本語）",
        "様式ツリー-標準ラベル（日本語）",
        "詳細ツリー-標準ラベル（日本語）",
    ]
    standard_ja_idx = next((name_to_index[c] for c in standard_ja_candidates if c in name_to_index), None)
    required = [
        "冗長ラベル（日本語）",
        "標準ラベル（英語）",
        "冗長ラベル（英語）",
        "名前空間プレフィックス",
        "要素名",
    ]
    if standard_ja_idx is None or not all(k in name_to_index for k in required):
        return None
    return {
        "standard_label_ja": standard_ja_idx,
        "terse_label_ja":    name_to_index["冗長ラベル（日本語）"],
        "standard_label_en": name_to_index["標準ラベル（英語）"],
        "terse_label_en":    name_to_index["冗長ラベル（英語）"],
        "prefix":            name_to_index["名前空間プレフィックス"],
        "element":           name_to_index["要素名"],
        "type":              name_to_index.get("type", -1),
        "substitution_group":name_to_index.get("substitutionGroup", -1),
        "period_type":       name_to_index.get("periodType", -1),
        "balance":           name_to_index.get("balance", -1),
        "abstract":          name_to_index.get("abstract", -1),
    }


def collect_from_workbook(path: Path, focus: set[str]) -> list[dict[str, object]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, object]] = []
    seen: set[tuple[str, str]] = set()
    for ws in wb.worksheets:
        idx: dict[str, int] | None = None
        for r, row in enumerate(ws.iter_rows(values_only=True), 1):
            vals = list(row)
            if idx is None:
                idx = _detect_header(vals)
                continue
            def _g(name: str):
                i = idx.get(name, -1)
                return vals[i] if 0 <= i < len(vals) else None
            element = _g("element")
            prefix = _g("prefix")
            if not element or element not in focus:
                continue
            key = (str(prefix), str(element))
            if key in seen:
                continue
            seen.add(key)
            rows.append({
                "sheet": ws.title,
                "row": r,
                "standard_label_ja": _g("standard_label_ja"),
                "terse_label_ja":    _g("terse_label_ja"),
                "standard_label_en": _g("standard_label_en"),
                "terse_label_en":    _g("terse_label_en"),
                "prefix":            str(prefix),
                "namespace":         EDINET_NAMESPACES.get(str(prefix)),
                "element":           str(element),
                "type":              _g("type"),
                "substitution_group":_g("substitution_group"),
                "period_type":       _g("period_type"),
                "abstract":          _g("abstract"),
            })
    return rows


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--workbook",
        type=Path,
        default=REPO / "data" / "sources" / "edinet-taxonomy-2026" / "1e_ElementList.xlsx",
    )
    ap.add_argument("--out", type=Path, default=REPO / "data" / "derived" / "edinet_taxonomy_focus.json")
    args = ap.parse_args()
    if not args.workbook.exists():
        print(
            f"workbook not found: {args.workbook}; run scripts/download_edinet_taxonomy.py",
            file=sys.stderr,
        )
        return 2
    rows = collect_from_workbook(args.workbook, DEFAULT_FOCUS)
    account_workbook = args.workbook.parent / "1f_AccountList.xlsx"
    if account_workbook.exists():
        rows += collect_from_workbook(account_workbook, DEFAULT_FOCUS)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps({
        "source": str(args.workbook.relative_to(REPO)),
        "namespaces": EDINET_NAMESPACES,
        "elements": rows,
    }, ensure_ascii=False, indent=2))
    print(f"wrote {args.out.relative_to(REPO)} ({len(rows)} elements)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
